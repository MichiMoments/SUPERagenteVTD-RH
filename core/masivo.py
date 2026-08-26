"""Carga masiva de otrosíes: plantilla de Excel, lectura de filas y .zip de .docx.

Sin dependencia de Streamlit, igual que documento.py: la interfaz solo entrega
bytes y muestra lo que devuelven estas funciones. Todas las funciones reciben el
tipo de otrosí, que es de donde salen las columnas, los formatos y las listas.
"""

import io
import re
import unicodedata
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.datavalidation import DataValidation

from . import campos
from . import documento
from . import tipos

HOJA_DATOS, HOJA_INSTRUCCIONES = "Otrosíes", "Instrucciones"
FILAS_PLANTILLA = 300
MAXIMO_FILAS = 300          # 42 ms medidos por documento: 300 filas son ~13 s de espera
FILAS_ENCABEZADO = 10       # hasta qué fila se busca el encabezado

# Nunca formato Texto («@») en las fechas: rompería justo lo que se quiere lograr.
FORMATOS = {"fecha": "DD/MM/YYYY", "cedula": "#,##0", "entero": "#,##0"}

# Primera columna de «Instrucciones» donde se copian las opciones de las listas; a cada
# campo de elección se le da la siguiente. Antes eran H e I a mano, con dos listas fijas.
COLUMNA_LISTAS = 8  # H

_ERROR_ARCHIVO = (
    "No pude abrir el archivo como libro de Excel. Ábrelo en Excel y usa «Guardar como → "
    "Libro de Excel (.xlsx)»; un .csv o un .xls renombrado no sirve. Detalle técnico: {}"
)


def construir_plantilla(tipo):
    """Arma el .xlsx vacío (hoja de datos + instrucciones) y devuelve los bytes."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = HOJA_DATOS
    # las instrucciones se escriben antes porque los desplegables apuntan a sus celdas
    _escribir_instrucciones(libro.create_sheet(HOJA_INSTRUCCIONES), tipo)
    _escribir_encabezado(hoja, tipo)
    _formatear_columnas(hoja, tipo)
    _agregar_desplegables(hoja, tipo)
    libro.active = 0  # que el libro abra en la hoja de datos y no en las instrucciones
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


_TECHO_FILAS = MAXIMO_FILAS + 20  # margen para filas vacías intermedias


def leer_libro(tipo, contenido, fecha_firma_defecto):
    """Lee el .xlsx cargado -> (registros, errores): un registro por fila con datos."""
    datos = io.BytesIO(contenido)

    # pre-chequeo liviano: read_only=True no materializa celdas, así que detectar un
    # rango inflado (alguien tocó la fila 500 000 en Excel) cuesta milisegundos
    try:
        rapido = load_workbook(datos, read_only=True)
    except (zipfile.BadZipFile, InvalidFileException, KeyError, ValueError) as error:
        return [], [_ERROR_ARCHIVO.format(error)]
    try:
        try:
            hoja_rapida = rapido[HOJA_DATOS]
        except KeyError:
            hoja_rapida = rapido.worksheets[0] if rapido.worksheets else None
        filas_declaradas = hoja_rapida.max_row if hoja_rapida else 0
    finally:
        rapido.close()

    if (filas_declaradas or 0) > _TECHO_FILAS:
        return [], [
            f"El archivo parece tener datos o formato hasta la fila {filas_declaradas}. "
            f"El tope es {MAXIMO_FILAS} filas de datos. Descarga una plantilla nueva "
            "o revisa que no haya celdas modificadas muy abajo en la hoja."
        ]

    datos.seek(0)
    try:
        libro = load_workbook(datos, data_only=True)
    except (zipfile.BadZipFile, InvalidFileException, KeyError, ValueError) as error:
        return [], [_ERROR_ARCHIVO.format(error)]
    try:
        try:
            hoja = _hoja_datos(libro, tipo)
            encabezado, mapa = _mapa_columnas(hoja, tipo)
        except ValueError as error:
            return [], [str(error)]

        del_lote = [
            campo["clave"]
            for campo in tipo["campos"]
            if campo["opcional_en_hoja"] and campo["tipo"] == "fecha"
        ]

        registros, excedido = [], False
        techo = min(hoja.max_row or encabezado, encabezado + MAXIMO_FILAS + 2)
        for numero in range(encabezado + 1, techo + 1):
            valores = {
                clave: hoja.cell(row=numero, column=columna).value
                for clave, columna in mapa.items()
            }
            if _fila_vacia(valores):
                continue
            if len(registros) >= MAXIMO_FILAS:
                excedido = True
                break
            for clave in del_lote:
                if valores.get(clave) in (None, ""):
                    valores[clave] = fecha_firma_defecto
            fila_datos, errores, avisos = campos.normalizar(tipo, valores)
            registros.append({
                "fila": numero,
                "datos": fila_datos,
                "errores": [f"Fila {numero} · {texto}" for texto in errores],
                "avisos": [f"Fila {numero} · {texto}" for texto in avisos],
            })
    finally:
        libro.close()

    _nombres_unicos(tipo, registros)
    _avisos_cruzados(tipo, registros)

    problemas = []
    if excedido:
        problemas.append(
            f"El archivo trae más de {MAXIMO_FILAS} filas con datos, que es el tope. "
            "Pártelo en varios archivos y súbelos uno por uno."
        )
    elif not registros:
        problemas.append(
            "Subiste la plantilla vacía: no encontré ninguna fila con datos debajo del "
            f"encabezado de la hoja «{hoja.title}»."
        )
    if registros and (encabezado != 1 or _llave(hoja.title) != _llave(HOJA_DATOS)):
        registros[0]["avisos"].insert(
            0, f"Leí los encabezados en la fila {encabezado} de la hoja «{hoja.title}»."
        )
    problemas.extend(error for registro in registros for error in registro["errores"])
    return registros, problemas


def generar_zip(tipo, registros, progreso=None):
    """Un .docx por registro, empaquetados -> (bytes, fallos, nombres realmente escritos)."""
    buffer = io.BytesIO()
    fallos, generados = [], []
    total = len(registros)
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as paquete:
        for hechos, registro in enumerate(registros, start=1):
            try:
                markdown = documento.render_markdown(tipo, registro["datos"])
                paquete.writestr(
                    registro["archivo"], documento.markdown_a_docx(markdown, tipo["titulo"])
                )
            except Exception as error:  # una fila que falle no puede matar el lote entero
                fallos.append(f"Fila {registro['fila']} · no se pudo generar: {error}")
            else:
                generados.append(registro["archivo"])
            if progreso:
                progreso(hechos, total)
    return buffer.getvalue(), fallos, generados


def _hoja_datos(libro, tipo):
    """Ubica la hoja por nombre y si no por encabezados; wb.active puede ser «Instrucciones»."""
    for hoja in libro.worksheets:
        if _llave(hoja.title) == _llave(HOJA_DATOS):
            return hoja
    # sin la hoja por nombre se toma la que más encabezados reconocibles tenga, y se deja
    # que _mapa_columnas dé el error preciso sobre ella en vez de un «no la encontré»
    mejor = max(libro.worksheets, key=lambda hoja: _reconocidos(hoja, tipo), default=None)
    if mejor is not None and _reconocidos(mejor, tipo):
        return mejor
    raise ValueError(
        f"No encontré la hoja «{HOJA_DATOS}» ni ninguna hoja con los encabezados de la "
        f"plantilla. Hojas del archivo: {', '.join(libro.sheetnames)}. Descarga la plantilla "
        "otra vez y pega tus datos en ella."
    )


def _reconocidos(hoja, tipo):
    """Cuántos encabezados de la plantilla aparecen en las primeras filas de la hoja."""
    esperadas = {_llave(campo["etiqueta"]) for campo in tipo["campos"]}
    vistas = set()
    for fila in range(1, min(FILAS_ENCABEZADO, hoja.max_row or 1) + 1):
        for columna in range(1, (hoja.max_column or 1) + 1):
            llave = _llave(hoja.cell(row=fila, column=columna).value or "")
            if llave in esperadas:
                vistas.add(llave)
    return len(vistas)


def _mapa_columnas(hoja, tipo):
    """Etiqueta -> índice de columna, tolerante a mayúsculas, tildes y espacios."""
    etiquetas = tipos.etiquetas(tipo)
    esperadas = {_llave(etiqueta): clave for clave, etiqueta in etiquetas.items()}
    minimo = max(1, len(etiquetas) // 2)  # con la mitad reconocida ya es la fila del encabezado
    for fila in range(1, min(FILAS_ENCABEZADO, hoja.max_row or 1) + 1):
        mapa, repetidas = {}, []
        for columna in range(1, (hoja.max_column or 1) + 1):
            clave = esperadas.get(_llave(hoja.cell(row=fila, column=columna).value or ""))
            if clave is None:
                continue
            if clave in mapa:
                repetidas.append(etiquetas[clave])
            else:
                mapa[clave] = columna
        if len(mapa) + len(repetidas) < minimo:
            continue
        if repetidas:
            raise ValueError(
                f"La fila {fila} de la hoja «{hoja.title}» tiene columnas repetidas: "
                + ", ".join(f"«{etiqueta}»" for etiqueta in sorted(set(repetidas)))
                + ". Deja una sola de cada una: si no, no hay forma de saber cuál vale."
            )
        if len(mapa) < len(etiquetas):
            faltan = [etiquetas[clave] for clave in etiquetas if clave not in mapa]
            raise ValueError(
                f"En la fila {fila} de la hoja «{hoja.title}» reconocí {len(mapa)} de "
                f"{len(etiquetas)} encabezados. Faltan: "
                + ", ".join(f"«{etiqueta}»" for etiqueta in faltan)
                + ". Ojo: una celda combinada deja el texto solo en la primera columna."
            )
        return fila, mapa
    raise ValueError(
        f"No encontré los encabezados de la plantilla en las primeras {FILAS_ENCABEZADO} "
        f"filas de la hoja «{hoja.title}». Descarga la plantilla y pega tus datos en ella."
    )


def _fila_vacia(valores):
    """True si todas las celdas mapeadas están vacías: relleno de la plantilla."""
    return all(
        valor is None or (isinstance(valor, str) and not valor.strip())
        for valor in valores.values()
    )


def _nombres_unicos(tipo, registros):
    """Nombres del .zip sin colisiones: dos grafías de un nombre dan el mismo slug."""
    usados = set()
    clave_nombre = tipo.get("campo_nombre")
    for registro in registros:
        if registro["errores"]:
            continue  # sin datos completos no hay nombre que calcular, y el lote ya está roto
        datos = registro["datos"]
        base = documento.nombre_archivo(tipo, datos)
        # el slug queda vacío si el nombre no deja ni una letra ASCII; se detecta comparando
        # con el nombre que saldría sin nombre, para no depender del formato de la plantilla
        if clave_nombre and base == documento.nombre_archivo(
            tipo, {**datos, clave_nombre: ""}
        ):
            base = documento.nombre_archivo(
                tipo, {**datos, clave_nombre: f"fila {registro['fila']}"}
            )
        raiz, _, extension = base.rpartition(".")
        nombre, repeticion = base, 1
        while nombre in usados:
            repeticion += 1
            nombre = f"{raiz}_{repeticion}.{extension}"
        usados.add(nombre)
        registro["archivo"] = nombre


def _avisos_cruzados(tipo, registros):
    """Avisos entre filas: documentos de identidad repetidos y archivos renombrados."""
    claves_cedula = [campo["clave"] for campo in tipo["campos"] if campo["tipo"] == "cedula"]
    for clave in claves_cedula:
        filas_por_cedula = {}
        for registro in registros:
            cedula = registro["datos"].get(clave)
            if cedula:
                filas_por_cedula.setdefault(cedula, []).append(registro["fila"])
        for registro in registros:
            filas = filas_por_cedula.get(registro["datos"].get(clave), [])
            if len(filas) > 1:
                registro["avisos"].append(
                    f"Fila {registro['fila']} · la cédula "
                    f"{documento.cedula(registro['datos'][clave])} está repetida "
                    f"(filas {', '.join(str(numero) for numero in filas)})"
                )
    for registro in registros:
        archivo = registro.get("archivo")
        if archivo and archivo != documento.nombre_archivo(tipo, registro["datos"]):
            registro["avisos"].append(
                f"Fila {registro['fila']} · el archivo se guarda como «{archivo}»: el nombre "
                "que le tocaba ya estaba usado o se quedaba sin letras"
            )


def _llave(texto):
    """Encabezado a clave de comparación: ' Fecha de INGRESO: ' -> 'fecha de ingreso'."""
    plano = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", " ", plano.lower()).strip()


def _escribir_encabezado(hoja, tipo):
    """Fila 1 con las etiquetas en el orden de los campos, en negrita y congelada."""
    relleno = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    for columna, campo in enumerate(tipo["campos"], start=1):
        celda = hoja.cell(row=1, column=columna, value=campo["etiqueta"])
        celda.font = Font(bold=True, color="FFFFFFFF")
        celda.fill = relleno
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    hoja.row_dimensions[1].height = 32
    hoja.freeze_panes = "A2"


def _formatear_columnas(hoja, tipo):
    """Anchos, y formato de fecha o de número en las columnas que lo necesitan."""
    for columna, campo in enumerate(tipo["campos"], start=1):
        hoja.column_dimensions[get_column_letter(columna)].width = campo["ancho"]
        formato = FORMATOS.get(campo["tipo"])
        if not formato:
            continue
        # celda por celda: dar formato crea las celdas, y por eso max_row será 301 aunque
        # la hoja esté vacía. _fila_vacia es lo que hace que eso no importe
        for fila in range(2, FILAS_PLANTILLA + 2):
            hoja.cell(row=fila, column=columna).number_format = formato


def _campos_lista(tipo):
    """[(campo, letra en «Instrucciones»)] de los campos de elección, en orden."""
    elegibles = [campo for campo in tipo["campos"] if tipos.opciones(campo)]
    return [
        (campo, get_column_letter(COLUMNA_LISTAS + indice))
        for indice, campo in enumerate(elegibles)
    ]


def _agregar_desplegables(hoja, tipo):
    """Listas de los campos de elección, apuntando a las celdas de «Instrucciones»."""
    posiciones = {campo["clave"]: indice for indice, campo in enumerate(tipo["campos"])}
    for campo, letra in _campos_lista(tipo):
        columna = get_column_letter(posiciones[campo["clave"]] + 1)
        opciones = tipos.opciones(campo)
        validacion = DataValidation(
            type="list",
            # el rango, y no una lista en línea («"Femenino,Masculino"»): el separador de una
            # lista en línea depende del idioma de Excel y, si no coincide, el desplegable no
            # aparece y nadie se da cuenta. De paso son la tabla de valores permitidos
            formula1=f"{HOJA_INSTRUCCIONES}!${letra}$2:${letra}${len(opciones) + 1}",
            allowBlank=True,
            showErrorMessage=True,
            errorTitle="Valor no válido",
            error="Elige uno de los valores de la lista desplegable.",
        )
        # showDropDown se deja sin poner a propósito: en OOXML el «1» oculta la flecha
        hoja.add_data_validation(validacion)
        validacion.add(f"{columna}2:{columna}{FILAS_PLANTILLA + 1}")


def _instrucciones(tipo):
    """Las líneas de «cómo diligenciar», omitiendo las que no apliquen a este tipo."""
    lineas = [
        "Una fila por persona, a partir de la fila 2.",
        "No traduzcas, muevas ni borres los encabezados de la fila 1. Puedes añadir columnas "
        "tuyas: se ignoran.",
    ]
    if any(campo["tipo"] == "fecha" for campo in tipo["campos"]):
        lineas.append(
            "Escribe las fechas en celdas con formato de fecha, nunca como texto. "
            "«03/04/2026» escrito como texto puede ser el 3 de abril o el 4 de marzo y no "
            "hay forma de saberlo, así que se rechaza. Como texto solo se acepta AAAA-MM-DD."
        )
    for campo in tipo["campos"]:
        if campo["opcional_en_hoja"]:
            lineas.append(
                f"«{campo['etiqueta']}» puede ir en blanco: la app le pone la del lote."
            )
    if elegibles := [campo for campo, _ in _campos_lista(tipo)]:
        lineas.append(
            "Usa los desplegables de "
            + ", ".join(f"«{campo['etiqueta']}»" for campo in elegibles)
            + ". Si pegas datos de otro archivo, Excel borra el desplegable: revisa que el "
            "texto quede idéntico."
        )
    lineas.append(
        "No uses «|» ni «**» en ningún campo: el documento se arma con tablas y negrita en "
        "Markdown, y esos caracteres borran en silencio lo que viene después."
    )
    if articulo := [campo for campo in tipo["campos"] if campo["articulo_minuscula"]]:
        lineas.append(
            ", ".join(f"«{campo['etiqueta']}»" for campo in articulo)
            + " va con su artículo en minúscula («la Dirección de…»): se imprime a mitad de "
            "frase. Y ningún campo puede empezar por «-»."
        )
    else:
        lineas.append("Ningún campo puede empezar por «-» ni por «|».")
    lineas.append(
        "Evita Alt+Enter dentro de una celda: el salto de línea se une en un solo renglón y "
        "la app te lo avisa."
    )
    if any(campo["tipo"] == "cedula" for campo in tipo["campos"]):
        lineas.append("Escribe la cédula solo con dígitos, sin «CC» y sin letras.")
    lineas.extend([
        "Si llenas celdas con fórmulas (BUSCARV y compañía), abre y guarda el archivo en "
        "Excel antes de subirlo: la app lee el último valor que Excel dejó calculado.",
        "Si el archivo abre en Vista Protegida, pulsa «Habilitar edición» o los desplegables "
        "no funcionarán.",
        f"Máximo {MAXIMO_FILAS} filas con datos por archivo. Si tienes más, pártelo en varios.",
    ])
    return [f"{numero}. {linea}" for numero, linea in enumerate(lineas, start=1)]


def _escribir_instrucciones(hoja, tipo):
    """Hoja 2: cómo diligenciar, y la tabla por campo generada desde el propio tipo."""
    hoja["A1"] = f"Cómo diligenciar la plantilla de «{tipo['nombre']}»"
    hoja["A1"].font = Font(bold=True, size=14)
    lineas = _instrucciones(tipo)
    for numero, linea in enumerate(lineas, start=3):
        hoja.cell(row=numero, column=1, value=linea).alignment = Alignment(wrap_text=True)

    fila = len(lineas) + 5
    for columna, titulo in enumerate(
        ("Campo", "Obligatorio", "Tipo", "Valores permitidos", "Ejemplo"), start=1
    ):
        hoja.cell(row=fila, column=columna, value=titulo).font = Font(bold=True)
    for campo in tipo["campos"]:
        fila += 1
        valores = (
            campo["etiqueta"],
            "No (la pone la app)" if campo["opcional_en_hoja"] else
            "Sí" if campo["obligatorio"] else "No",
            tipos.NOMBRE_TIPO_CAMPO.get(campo["tipo"], campo["tipo"]),
            _permitidos(campo),
            "(en blanco)" if campo["opcional_en_hoja"] else _ejemplo(campo),
        )
        for columna, valor in enumerate(valores, start=1):
            hoja.cell(row=fila, column=columna, value=valor)
    for columna, ancho in zip("ABCDE", (38, 20, 14, 48, 46)):
        hoja.column_dimensions[columna].width = ancho

    listas = _campos_lista(tipo)
    if listas:
        hoja.cell(row=1, column=COLUMNA_LISTAS,
                  value="No borrar: son los valores de las listas desplegables").font = Font(bold=True)
    for campo, letra in listas:
        for indice, opcion in enumerate(tipos.opciones(campo), start=2):
            hoja[f"{letra}{indice}"] = opcion
        hoja.column_dimensions[letra].width = 26


def _ejemplo(campo):
    """El ejemplo del campo tal como se lee: la fecha ISO del .json se muestra DD/MM/AAAA."""
    ejemplo = campo["ejemplo"]
    if campo["tipo"] == "fecha" and ejemplo:
        try:
            return f"{campos._fecha(ejemplo):%d/%m/%Y}"
        except ValueError:
            return str(ejemplo)
    return str(ejemplo)


def _permitidos(campo):
    """Valores que acepta una columna: 'lista' -> 'Dos (2) días… / Tres (3) días…'."""
    if campo["tipo"] == "fecha":
        return "Fecha (celda con formato de fecha); como texto, solo AAAA-MM-DD"
    if campo["tipo"] in ("cedula", "entero"):
        return "Solo dígitos, sin «CC» ni letras ni notación científica"
    if opciones := tipos.opciones(campo):
        return " / ".join(opciones)
    return "Texto, sin «|» ni «**», y sin empezar por «-»"
